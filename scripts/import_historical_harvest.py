#!/usr/bin/env python3
"""One-off import of pre-app historical harvest data (2020-2025) into the
HistoricalHarvest table, powering the admin Analysis tab.

Source: data/imports/"Harvest Data 2020-2025 (Clean, Split by Block).xlsx"
(gitignored - per-farm data, not shipped with the app)
(itself derived from the farm's original "Daaglikse Oes data 2020 - 2025.xlsx" -
see that clean workbook's own Notes sheet for the block-split-by-hectare-ratio
and column-typo caveats that carry through to the season_year/block_id/
estimated fields below).

Safe to re-run: replaces the whole table each time, so re-running after
regenerating the clean workbook (e.g. the block register changed) just
reloads it with the new numbers.

Usage (run with the backend's own venv so sqlmodel etc. are on the path):
    backend/.venv/bin/python3 scripts/import_historical_harvest.py
"""
import os
import sys

import openpyxl

BACKEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend")
sys.path.insert(0, BACKEND_DIR)

from sqlmodel import Session, delete  # noqa: E402

from db import create_db_and_tables, engine  # noqa: E402
from models import HistoricalHarvest  # noqa: E402

# The source workbook lives in data/imports/, which is gitignored, NOT in the
# repository. It used to be committed - which meant every install carried one
# particular farm's harvest records, and a new customer's server would clone
# down another farm's business data and import it. data/ is where per-farm
# material belongs; the repository ships code.
#
# Pass a different path as the first argument if the workbook is elsewhere.
DEFAULT_XLSX = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data", "imports",
                          "Harvest Data 2020-2025 (Clean, Split by Block).xlsx")
XLSX_PATH = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
YEAR_SHEETS = ["2020", "2021", "2022", "2023", "2024", "2025"]
SPLIT_BLOCKS = {"8a", "8b", "10a", "10b", "17a", "17b", "19a", "19b"}


def load_rows():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    rows = []
    for year in YEAR_SHEETS:
        ws = wb[year]
        header = [c.value for c in ws[1]]
        block_cols = [(i, h) for i, h in enumerate(header) if h not in (None, "Date", "Weekday", "Total")]
        for row in ws.iter_rows(min_row=2, values_only=True):
            harvest_date = row[0]
            if harvest_date is None or not hasattr(harvest_date, "year"):
                continue  # the TOTAL footer row, or a blank trailing row
            for idx, block_id in block_cols:
                kg = row[idx]
                if kg is None:
                    continue
                block_id = str(block_id)
                rows.append(HistoricalHarvest(
                    block_id=block_id, harvest_date=harvest_date.date(), season_year=int(year),
                    kg=float(kg), estimated=block_id in SPLIT_BLOCKS,
                ))
    return rows


def main():
    if not os.path.exists(XLSX_PATH):
        print(f"No source workbook at {XLSX_PATH} - skipping the daily harvest import.\n"
              "This is normal on a new install: the workbook holds one farm's own\n"
              "historical records and is not shipped with the app. Put yours in\n"
              "data/imports/, or pass its path as the first argument, then re-run.\n"
              "Nothing was changed.")
        return
    create_db_and_tables()
    records = load_rows()
    with Session(engine) as session:
        session.exec(delete(HistoricalHarvest))
        session.add_all(records)
        session.commit()
    print(f"Imported {len(records)} historical harvest rows from:\n  {XLSX_PATH}")


if __name__ == "__main__":
    main()
