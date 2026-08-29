#!/usr/bin/env python3
"""Self-test for the historical-data, Risk indicator and Harvest Forecast
work - the parts of this app with real arithmetic behind them, where a
wrong answer looks just as plausible as a right one.

Runs against the server's OWN database, read-only: it never writes, so it
is safe to run on the live farm server to confirm an install is sound.
Anything that depends on the farm's specific numbers is asserted as an
invariant ("the sheets reconcile", "every season is scored") rather than a
hardcoded figure, so this keeps working as seasons are added.

Plain asserts, no pytest - this project deliberately carries no test
dependency (same reason routers/risk.py hand-rolls its own least-squares
fit instead of pulling in numpy).

Usage:
    backend/.venv/bin/python3 scripts/selftest.py
Exits non-zero if anything fails.
"""
import io
import os
import shutil
import sqlite3
import sys
import tempfile
from collections import defaultdict
from datetime import date, datetime, timedelta
from types import SimpleNamespace

import openpyxl

BACKEND_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend")
sys.path.insert(0, BACKEND_DIR)

from alembic import command  # noqa: E402
from alembic.autogenerate import compare_metadata  # noqa: E402
from alembic.migration import MigrationContext  # noqa: E402
from alembic.script import ScriptDirectory  # noqa: E402
from sqlalchemy import create_engine, inspect  # noqa: E402
from sqlmodel import Session, SQLModel, func, select  # noqa: E402

import asyncio  # noqa: E402
from fastapi import HTTPException, UploadFile  # noqa: E402

import backup  # noqa: E402
from db import DB_PATH, engine, legacy_schema_catch_up  # noqa: E402
from routers.master_data import import_blocks  # noqa: E402
from migrate import (BASELINE_REVISION, _baseline_database, _config,  # noqa: E402
                     current_revision, head_revision, run_migrations)
from models import (AdminUser, Block, HarvestRecord, HistoricalAnnualYield,  # noqa: E402
                    HistoricalHarvest, SetupState, SystemSetting, Worker, WeatherHistory)
from routers.analysis import build_analysis_summary  # noqa: E402
from routers.setup import build_setup_state  # noqa: E402
from routers.reports import historical_harvest_data_report  # noqa: E402
from routers.risk import (DRIVERS, REFERENCE_START_YEAR,  # noqa: E402
                          REGRESSION_START_YEAR, _band, _compute_driver_state,
                          _driver_value, _ols_fit, _project_driver, _reference_label,
                          _risk_points, _segment_days, _window_status,
                          build_harvest_forecast, build_risk_summary)
from routers.weather import build_weather_history  # noqa: E402

import routers.risk as risk_module  # noqa: E402

_passed = 0
_failed = []
_skipped = []


class Skip(Exception):
    """Raised by a check that cannot run against this particular database.

    Only for a genuinely absent precondition - a table the server has not
    created yet, say - never for an assertion that is inconvenient. Skips
    are counted and listed separately so they can't be mistaken for passes.
    """


class offline:
    """Runs the forecast without touching the network or the database.

    build_harvest_forecast() normally calls sync_recent_weather() (which
    APPENDS rows - this suite promises to be read-only, so it must not run
    here) and fetches a live Open-Meteo forecast (slow, and it would make
    results depend on today's weather). Both are stubbed: the forecast
    fetch raises, which is the same graceful-degradation path a farm
    server takes when its internet is down, so these tests exercise that
    fallback rather than skipping it."""

    def __enter__(self):
        self._sync = risk_module.sync_recent_weather
        self._fetch = risk_module.fetch_forecast_hourly
        risk_module.sync_recent_weather = lambda session: {"synced": 0}

        def _no_network(*a, **kw):
            raise OSError("network disabled for self-test")
        risk_module.fetch_forecast_hourly = _no_network
        return self

    def __exit__(self, *exc):
        risk_module.sync_recent_weather = self._sync
        risk_module.fetch_forecast_hourly = self._fetch
        return False


def check(name, fn):
    global _passed
    try:
        fn()
    except Skip as exc:
        _skipped.append((name, str(exc)))
        print(f"  skip  {name}\n          {exc}")
    except Exception as exc:  # noqa: BLE001 - a failing check must not stop the run
        _failed.append((name, exc))
        print(f"  FAIL  {name}\n          {type(exc).__name__}: {exc}")
    else:
        _passed += 1
        print(f"  ok    {name}")


def section(title):
    print(f"\n{title}\n{'-' * len(title)}")


# ---------------------------------------------------------------------------
# Pure-function tests - no database, exact expected values
# ---------------------------------------------------------------------------
def test_risk_points_scaling():
    hist = [10.0, 20.0]
    # lower_is_worse: the low end is maximum risk
    assert _risk_points(10.0, hist, "lower_is_worse") == 25.0
    assert _risk_points(20.0, hist, "lower_is_worse") == 0.0
    assert _risk_points(15.0, hist, "lower_is_worse") == 12.5
    # higher_is_worse inverts it
    assert _risk_points(20.0, hist, "higher_is_worse") == 25.0
    assert _risk_points(10.0, hist, "higher_is_worse") == 0.0


def test_risk_points_clamps_beyond_history():
    hist = [10.0, 20.0]
    # Past either extreme must saturate, never run off the 0-25 scale
    assert _risk_points(5.0, hist, "lower_is_worse") == 25.0
    assert _risk_points(99.0, hist, "lower_is_worse") == 0.0
    assert _risk_points(99.0, hist, "higher_is_worse") == 25.0
    assert _risk_points(-99.0, hist, "higher_is_worse") == 0.0


def test_risk_points_degenerate():
    assert _risk_points(None, [1.0, 2.0], "lower_is_worse") is None
    assert _risk_points(1.0, [], "lower_is_worse") is None
    assert _risk_points(5.0, [5.0, 5.0], "lower_is_worse") == 12.5  # no spread -> neutral


def test_band_boundaries():
    assert _band(0) == "Low"
    assert _band(24.9) == "Low"
    assert _band(25) == "Moderate"
    assert _band(49.9) == "Moderate"
    assert _band(50) == "Elevated"
    assert _band(74.9) == "Elevated"
    assert _band(75) == "High"
    assert _band(100) == "High"


def test_driver_value_aggregations():
    def row(ts, **kw):
        return SimpleNamespace(timestamp=ts, **kw)
    day = datetime(2024, 10, 1)
    rows = [row(day + timedelta(hours=h), temp_c=float(h), precipitation_mm=1.0,
                sunshine_duration_s=3600.0) for h in range(24)]
    rows += [row(day + timedelta(days=1, hours=h), temp_c=float(h) + 10,
                 precipitation_mm=0.0, sunshine_duration_s=1800.0) for h in range(24)]

    assert _driver_value(rows, {"field": "temp_c", "agg": "count_lt", "threshold": 5}) == 5.0
    assert _driver_value(rows, {"field": "temp_c", "agg": "count_gt", "threshold": 30}) == 3.0
    # mean of 0..23 and 10..33 = 11.5 and 21.5 -> 16.5
    assert _driver_value(rows, {"field": "temp_c", "agg": "mean"}) == 16.5
    # daily maxima 23 and 33 -> 28
    assert _driver_value(rows, {"field": "temp_c", "agg": "daily_max_mean"}) == 28.0
    assert _driver_value(rows, {"field": "precipitation_mm", "agg": "sum"}) == 24.0
    assert _driver_value(rows, {"field": "precipitation_mm", "agg": "count_days_gt",
                                 "threshold": 0.5}) == 1.0
    # scale converts seconds -> hours: 24*3600 + 24*1800 = 129600s = 36h
    assert _driver_value(rows, {"field": "sunshine_duration_s", "agg": "sum",
                                 "scale": 1 / 3600}) == 36.0


def test_driver_value_missing_data():
    def row(ts, v):
        return SimpleNamespace(timestamp=ts, temp_c=v)
    day = datetime(2024, 10, 1)
    assert _driver_value([], {"field": "temp_c", "agg": "mean"}) is None
    allnull = [row(day, None), row(day, None)]
    assert _driver_value(allnull, {"field": "temp_c", "agg": "mean"}) is None
    assert _driver_value(allnull, {"field": "temp_c", "agg": "daily_max_mean"}) is None
    assert _driver_value(allnull, {"field": "temp_c", "agg": "count_lt", "threshold": 5}) is None
    # a partial gap still yields a value from the hours that do have data
    mixed = [row(day, None), row(day, 10.0)]
    assert _driver_value(mixed, {"field": "temp_c", "agg": "mean"}) == 10.0


def test_every_driver_agg_is_implemented():
    """A driver added with an unimplemented agg must not reach production -
    _driver_value raises on unknown aggs, so exercise each one."""
    def row(ts, **kw):
        return SimpleNamespace(timestamp=ts, **kw)
    day = datetime(2024, 10, 1)
    for d in DRIVERS:
        rows = [row(day + timedelta(hours=h), **{d["field"]: 1.0}) for h in range(3)]
        _driver_value(rows, d)  # raises ValueError on an unknown agg


def test_window_status_transitions():
    wmd = ((9, 16), (10, 31))
    assert _window_status(2025, wmd, date(2025, 9, 15)) == "pending"
    assert _window_status(2025, wmd, date(2025, 9, 16)) == "in_progress"
    assert _window_status(2025, wmd, date(2025, 10, 31)) == "in_progress"
    assert _window_status(2025, wmd, date(2025, 11, 1)) == "final"


def test_segment_days_partition():
    """actual/forecast/assumed must tile the window exactly - no gaps, no
    double counting - for every position of "today" relative to it."""
    ws, we = date(2025, 10, 1), date(2025, 10, 30)  # 30 days
    total = (we - ws).days + 1
    for offset in range(-5, 36):
        today = ws + timedelta(days=offset)
        for horizon in (0, 3, 15, 60):
            segs = _segment_days(ws, we, today, horizon)
            days = 0
            seen = set()
            for seg in segs.values():
                if seg is None:
                    continue
                assert seg[0] <= seg[1], (seg, today, horizon)
                assert ws <= seg[0] and seg[1] <= we, (seg, today, horizon)
                d = seg[0]
                while d <= seg[1]:
                    assert d not in seen, f"day {d} counted twice"
                    seen.add(d)
                    d += timedelta(days=1)
                    days += 1
            if today >= we:            # window fully behind us
                assert days == total, (days, total, today, horizon)
            if today < ws and horizon == 0:  # entirely ahead, no forecast
                assert segs["assumed"] == (ws, we)


def test_segment_days_zero_horizon_has_no_forecast():
    segs = _segment_days(date(2025, 10, 1), date(2025, 10, 30), date(2025, 10, 10), 0)
    assert segs["forecast"] is None
    assert segs["actual"] == (date(2025, 10, 1), date(2025, 10, 10))
    assert segs["assumed"] == (date(2025, 10, 11), date(2025, 10, 30))


def test_ols_fit():
    fit = _ols_fit([0.0, 1.0, 2.0, 3.0], [1.0, 3.0, 5.0, 7.0])  # y = 2x + 1 exactly
    assert abs(fit["slope"] - 2.0) < 1e-6
    assert abs(fit["intercept"] - 1.0) < 1e-6
    assert abs(fit["r"] - 1.0) < 1e-6
    assert fit["n_seasons"] == 4
    assert _ols_fit([1.0], [1.0]) is None            # too few points
    assert _ols_fit([1.0, 1.0], [1.0, 2.0]) is None  # no variation in x


def test_reference_label():
    assert _reference_label([2012, 2013, 2025]) == "2012-2025"
    assert _reference_label([2019]) == "2019"
    assert _reference_label([]) == ""


def test_project_driver_intensive_vs_extensive():
    """A 10-day window, half elapsed. An intensive driver (a mean) must
    blend by day-weight; an extensive one (a running total) must accumulate
    the scenario's daily rate over the days still to come."""
    state = {
        "current_year": 2025, "today": date(2025, 10, 5),
        "by_date": defaultdict(list), "hist_range": {"k": [100.0, 200.0]},
    }
    for d in range(1, 6):  # 1-5 Oct actual: 10 units/day
        for h in range(24):
            state["by_date"][date(2025, 10, d)].append(
                SimpleNamespace(timestamp=datetime(2025, 10, d, h), v=10.0 / 24))

    ext = {"key": "k", "window_md": ((10, 1), (10, 10)), "field": "v", "agg": "sum",
           "direction": "lower_is_worse"}
    got = _project_driver(ext, state, defaultdict(list), 0)
    assert got["actual_days"] == 5 and got["assumed_days"] == 5
    # actual 50 + expected-scenario rate (150/10=15/day) over 5 remaining days
    assert abs(got["scenarios"]["expected"] - (50.0 + 15.0 * 5)) < 1e-6

    inten = dict(ext, agg="mean")
    got = _project_driver(inten, state, defaultdict(list), 0)
    actual_mean = 10.0 / 24
    assert abs(got["scenarios"]["expected"] - (5 * actual_mean + 5 * 150.0) / 10) < 1e-6


def test_project_driver_data_gap_falls_back():
    """A window with elapsed days but no data must fall back to the pure
    historical scenario, not treat the hole as zero risk."""
    state = {"current_year": 2025, "today": date(2025, 10, 5),
             "by_date": defaultdict(list), "hist_range": {"k": [100.0, 200.0]}}
    d = {"key": "k", "window_md": ((10, 1), (10, 10)), "field": "v", "agg": "sum",
         "direction": "lower_is_worse"}
    got = _project_driver(d, state, defaultdict(list), 0)
    assert got["data_gap"] is True
    assert got["assumed_days"] == 10 and got["actual_days"] == 0
    assert got["scenarios"]["expected"] == 150.0


# ---------------------------------------------------------------------------
# Configuration sanity - catches a mis-specified driver before it ships
# ---------------------------------------------------------------------------
def test_driver_definitions_wellformed():
    seen = set()
    for d in DRIVERS:
        for key in ("key", "label", "window_md", "window_label", "field", "agg",
                    "unit", "direction", "why"):
            assert key in d, f"{d.get('key')} missing {key}"
        assert d["key"] not in seen, f"duplicate driver key {d['key']}"
        seen.add(d["key"])
        assert d["direction"] in ("lower_is_worse", "higher_is_worse")
        (sm, sd), (em, ed) = d["window_md"]
        assert date(2024, sm, sd) <= date(2024, em, ed), f"{d['key']} window ends before it starts"
        # The whole design assumes windows never straddle a year boundary
        # (_compute_driver_state's min(all_years) weather floor relies on it)
        assert sm <= em, f"{d['key']} window spans a calendar year"
        assert hasattr(WeatherHistory, d["field"]), f"{d['key']} reads unknown field {d['field']}"


def test_driver_fields_are_fetchable_for_all_eras():
    """Every driver field must be one the archive backfill actually carries.
    soil_temperature_6cm and uv_index come back all-null before 2020, so a
    driver depending on them would silently have no reference values for
    the older seasons."""
    from weather import ARCHIVE_HOURLY_FIELDS, HOURLY_FIELDS
    api_name = {"temp_c": "temperature_2m", "humidity_pct": "relative_humidity_2m",
                "dew_point_c": "dew_point_2m", "precipitation_mm": "precipitation",
                "wind_speed_kmh": "wind_speed_10m", "soil_temp_6cm_c": "soil_temperature_6cm",
                "uv_index": "uv_index", "sunshine_duration_s": "sunshine_duration"}
    for d in DRIVERS:
        name = api_name[d["field"]]
        assert name in ARCHIVE_HOURLY_FIELDS, f"{d['key']} uses {name}, absent from the archive backfill"
        assert name in HOURLY_FIELDS, f"{d['key']} uses {name}, absent from the live/forecast fetch"


def test_regression_start_not_before_reference_start():
    assert REGRESSION_START_YEAR >= REFERENCE_START_YEAR


# Physically possible bounds per stored field, independent of this farm.
# Their real job is catching a UNIT error: sunshine is stored in seconds and
# reported in hours, and a dropped scale factor leaves the score itself
# correct (min-max normalizing cancels any linear scale) while the figure on
# screen reads in the millions.
_FIELD_BOUNDS = {
    "temp_c": (-20.0, 55.0),
    "dew_point_c": (-30.0, 35.0),
    "humidity_pct": (0.0, 100.0),
    "wind_speed_kmh": (0.0, 300.0),
    "precipitation_mm": (0.0, 5000.0),
    "sunshine_duration_s": (0.0, 24.0),  # per day, once scaled to hours
}


def test_driver_values_are_physically_plausible():
    with Session(engine) as s:
        state = _compute_driver_state(s)
    for d in DRIVERS:
        lo, hi = _FIELD_BOUNDS[d["field"]]
        start, end = date(2024, *d["window_md"][0]), date(2024, *d["window_md"][1])
        window_days = (end - start).days + 1
        # A running total accumulates over the window; an average does not.
        ceiling = hi * window_days if d["agg"] in ("sum",) else hi
        floor = lo * window_days if d["agg"] in ("sum",) and lo > 0 else lo
        for year in state["historical_years"]:
            v = state["value_by_year"][d["key"]][year]
            assert floor <= v <= ceiling, (
                f"{d['key']} = {v} in {year} is outside the possible range "
                f"{floor}..{ceiling} for '{d['unit']}' over {window_days} days "
                f"- a unit/scale error?")


# Aggregations that are an average per unit of time, named here independently
# of risk.py's own _INTENSIVE_AGGS so that widening or narrowing that set
# without thinking is caught rather than silently accepted.
_AVERAGING_AGGS = {"mean", "daily_max_mean"}


def test_averaging_drivers_blend_by_day_weight():
    """For an averaging driver, if the elapsed part of the window already
    reads exactly the scenario value, the projection must still read that
    value - blending it as a running total instead inflates it by however
    much of the window is left."""
    for d in DRIVERS:
        if d["agg"] not in _AVERAGING_AGGS:
            continue
        assert d["agg"] in risk_module._INTENSIVE_AGGS, (
            f"{d['key']} averages over time but is not in _INTENSIVE_AGGS, so a "
            f"partly-elapsed window will be blended as if it accumulated")
        (sm, sd), (em, ed) = d["window_md"]
        start, end = date(2025, sm, sd), date(2025, em, ed)
        midpoint = start + timedelta(days=(end - start).days // 2)
        value = 20.0
        by_date = defaultdict(list)
        day = start
        while day <= midpoint:
            for h in range(24):
                by_date[day].append(SimpleNamespace(
                    timestamp=datetime(day.year, day.month, day.day, h),
                    **{d["field"]: value}))
            day += timedelta(days=1)
        state = {"current_year": 2025, "today": midpoint, "by_date": by_date,
                 "hist_range": {d["key"]: [value, value]}}
        got = _project_driver(d, state, defaultdict(list), 0)
        assert got["data_gap"] is False, d["key"]
        for scenario, projected in got["scenarios"].items():
            assert abs(projected - value) < 1e-6, (
                f"{d['key']} {scenario}: projected {projected}, expected {value}")


def test_regression_seasons_all_had_a_bearing_orchard():
    """REGRESSION_START_YEAR exists to keep seasons where the replanted
    orchard had not yet come into bearing out of the kg fit - their small
    totals reflect young trees, not weather. Assert that directly against
    the data, so moving the constant back is caught even though the
    forecast still produces plausible-looking numbers."""
    with Session(engine) as s:
        state = _compute_driver_state(s)
        blocks_by_year = defaultdict(set)
        for a in s.exec(select(HistoricalAnnualYield)).all():
            if a.block_id is not None and a.kg:
                blocks_by_year[a.season_year].add(a.block_id)
        for h in s.exec(select(HistoricalHarvest)).all():
            if h.block_id is not None and h.kg:
                blocks_by_year[h.season_year].add(h.block_id)
    fitted = [y for y in state["historical_years"] if y >= REGRESSION_START_YEAR]
    assert fitted, "no seasons left to fit the kg regression on"
    for year in fitted:
        assert len(blocks_by_year[year]) > 1, (
            f"{year} is in the kg regression but only "
            f"{len(blocks_by_year[year])} block(s) bore fruit - that total "
            f"reflects an immature orchard, not its weather")


# ---------------------------------------------------------------------------
# Database-backed integration - invariants, not hardcoded farm figures
# ---------------------------------------------------------------------------
def test_weather_history_continuous():
    """Gaps in the hourly record would silently bias any window average."""
    with Session(engine) as s:
        lo = s.exec(select(func.min(WeatherHistory.timestamp))).one()
        hi = s.exec(select(func.max(WeatherHistory.timestamp))).one()
        n = s.exec(select(func.count()).select_from(WeatherHistory)).one()
    assert lo is not None, "no weather history imported"
    expected = int((hi - lo).total_seconds() // 3600) + 1
    missing = expected - n
    assert missing <= 24, f"{missing} hourly rows missing between {lo} and {hi}"


def test_weather_covers_every_reference_window():
    with Session(engine) as s:
        state = _compute_driver_state(s)
    for d in DRIVERS:
        for year in state["historical_years"]:
            v = state["value_by_year"][d["key"]][year]
            assert v is not None, f"{d['key']} has no weather data for reference season {year}"


def test_no_duplicate_weather_hours():
    with Session(engine) as s:
        n = s.exec(select(func.count()).select_from(WeatherHistory)).one()
        distinct = s.exec(select(func.count(func.distinct(WeatherHistory.timestamp)))).one()
    assert n == distinct, f"{n - distinct} duplicate weather timestamps"


def test_historical_tables_dont_overlap():
    """HistoricalHarvest (daily) and HistoricalAnnualYield (season totals)
    must never cover the same season, or every total counts it twice."""
    with Session(engine) as s:
        daily = set(s.exec(select(HistoricalHarvest.season_year).distinct()).all())
        annual = set(s.exec(select(HistoricalAnnualYield.season_year).distinct()).all())
    assert not (daily & annual), f"seasons in both harvest tables: {sorted(daily & annual)}"


def test_annual_yield_block_ids_resolve():
    """Every block-level historical row must point at a real block, or the
    report's block columns silently drop it."""
    with Session(engine) as s:
        ids = {b.id for b in s.exec(select(Block)).all()}
        bad = {a.block_id for a in s.exec(select(HistoricalAnnualYield)).all()
               if a.block_id is not None and a.block_id not in ids}
        bad |= {h.block_id for h in s.exec(select(HistoricalHarvest)).all()
                if h.block_id is not None and h.block_id not in ids}
    assert not bad, f"harvest rows reference unknown blocks: {sorted(bad)}"


def test_risk_summary_every_reference_season_scored():
    with Session(engine) as s:
        r = build_risk_summary(s)
    assert r["driver_count"] == len(DRIVERS)
    for season in r["seasons"]:
        if season["year"] == r["current_year"]:
            continue
        assert season["risk_score"] is not None, f"{season['year']} has no score"
        assert 0 <= season["risk_score"] <= 100, season
        assert season["known_count"] == len(DRIVERS)
        assert season["band"] == _band(season["risk_score"])


def test_risk_component_points_sum_to_score():
    with Session(engine) as s:
        r = build_risk_summary(s)
    for season in r["seasons"]:
        pts = [c["risk_points"] for c in season["components"] if c["risk_points"] is not None]
        if season["risk_score"] is None:
            continue
        assert abs(sum(pts) - season["risk_score"]) < 0.05, season["year"]
        for c in season["components"]:
            if c["risk_points"] is not None:
                assert 0 <= c["risk_points"] <= 25, c


def test_risk_reference_extremes_hit_the_rails():
    """Each driver's own best and worst reference season must score exactly
    0 and 25 - if not, the normalization range isn't what it claims."""
    with Session(engine) as s:
        r = build_risk_summary(s)
        state = _compute_driver_state(s)
    for d in DRIVERS:
        vals = [state["value_by_year"][d["key"]][y] for y in state["historical_years"]]
        pts = [_risk_points(v, state["hist_range"][d["key"]], d["direction"]) for v in vals]
        assert min(pts) == 0.0 and max(pts) == 25.0, f"{d['key']} range not spanned: {pts}"
    for d in r["drivers"]:
        assert d["historical_min"] <= d["historical_max"]


def test_forecast_scenarios_ordered_and_bounded():
    with offline(), Session(engine) as s:
        f = build_harvest_forecast(s)
    sc = f["scenarios"]
    if sc["expected"]["predicted_kg"] is None:
        return  # no regression fitted (e.g. a fresh database) - nothing to order
    fav, exp, unf = sc["favorable"], sc["expected"], sc["unfavorable"]
    assert fav["risk_score"] <= exp["risk_score"] <= unf["risk_score"], sc
    # Lower risk must never predict less fruit
    assert fav["predicted_kg"] >= exp["predicted_kg"] >= unf["predicted_kg"], sc
    for s_ in (fav, exp, unf):
        assert s_["predicted_kg"] > 0, f"non-positive kg prediction: {s_}"
        assert 0 <= s_["risk_score"] <= 100


def test_forecast_predictions_within_recorded_range():
    """Predictions are clamped to the fitted seasons' own kg span - an
    unclamped line ran the unfavorable end past zero into negative kg."""
    with offline(), Session(engine) as s:
        f = build_harvest_forecast(s)
        state = _compute_driver_state(s)
    if f["scenarios"]["expected"]["predicted_kg"] is None:
        return
    kgs = [state["kg_by_year"][y] for y in state["historical_years"]
           if y >= REGRESSION_START_YEAR and state["kg_by_year"].get(y)]
    lo, hi = min(kgs), max(kgs)
    for name, sc in f["scenarios"].items():
        assert lo - 0.5 <= sc["predicted_kg"] <= hi + 0.5, f"{name} outside recorded range: {sc}"


def test_forecast_labels_match_their_own_ranges():
    """reference_label describes the normalization range, regression_label
    the (narrower) fitted one - mixing them credits the average to seasons
    it was never computed over."""
    with offline(), Session(engine) as s:
        f = build_harvest_forecast(s)
        state = _compute_driver_state(s)
    assert f["reference_label"] == _reference_label(state["historical_years"])
    reg_years = [y for y in state["historical_years"] if y >= REGRESSION_START_YEAR]
    assert f["regression_label"] == _reference_label(reg_years)
    assert f["reference_season_count"] == len(reg_years)
    if f["regression"]:
        assert f["regression"]["n_seasons"] == len(reg_years)


def test_reference_years_are_modern_orchard_only():
    with Session(engine) as s:
        state = _compute_driver_state(s)
    assert state["historical_years"], "no reference seasons at all"
    assert min(state["historical_years"]) >= REFERENCE_START_YEAR
    assert state["current_year"] not in state["historical_years"]


def test_risk_and_forecast_agree_on_shared_state():
    """Both features read one _compute_driver_state(); a season's score in
    the Risk tab must be the same number the forecast regression fitted."""
    with Session(engine) as s:
        r = build_risk_summary(s)
        state = _compute_driver_state(s)
    by_year = {x["year"]: x["risk_score"] for x in r["seasons"]}
    for year in state["historical_years"]:
        recomputed = sum(
            _risk_points(state["value_by_year"][d["key"]][year],
                          state["hist_range"][d["key"]], d["direction"])
            for d in DRIVERS)
        assert abs(recomputed - by_year[year]) < 0.05, year


def test_forecast_degrades_without_network():
    """A farm server with no internet must still render the card, falling
    back to pure historical scenarios rather than failing the request."""
    with offline(), Session(engine) as s:
        f = build_harvest_forecast(s)
    assert f["forecast_unavailable"] is True
    assert f["forecast_horizon_end"] is None
    assert f["scenarios"]["expected"]["risk_score"] is not None, \
        "forecast collapsed entirely when the provider was unreachable"
    for d in f["drivers"]:
        assert d["forecast_days"] == 0, d


def test_selftest_writes_nothing():
    """This suite is advertised as safe to run on the live server."""
    with Session(engine) as s:
        before = (s.exec(select(func.count()).select_from(WeatherHistory)).one(),
                  s.exec(select(func.count()).select_from(HarvestRecord)).one(),
                  s.exec(select(func.count()).select_from(HistoricalHarvest)).one(),
                  s.exec(select(func.count()).select_from(HistoricalAnnualYield)).one())
    with offline(), Session(engine) as s:
        build_risk_summary(s)
        build_harvest_forecast(s)
        historical_harvest_data_report(session=s, admin=None)
    with Session(engine) as s:
        after = (s.exec(select(func.count()).select_from(WeatherHistory)).one(),
                 s.exec(select(func.count()).select_from(HarvestRecord)).one(),
                 s.exec(select(func.count()).select_from(HistoricalHarvest)).one(),
                 s.exec(select(func.count()).select_from(HistoricalAnnualYield)).one())
    assert before == after, f"row counts changed: {before} -> {after}"


# ---------------------------------------------------------------------------
# Schema migrations
#
# Two kinds of check here. The ones that build a database do it in a
# temporary directory and throw it away - this suite is read-only against
# the server's own database and stays that way. The last two look at the
# live database, and only read.
#
# Worth having because the migration path is the one piece of this app that
# runs against a farm's only copy of its data, at the exact moment a release
# changes what that data looks like, on a machine nobody is watching.
# ---------------------------------------------------------------------------
def _table_shape(path):
    """The structure of every table in a SQLite file.

    Compared as structure rather than as CREATE TABLE text on purpose:
    Alembic emits foreign keys alphabetically and SQLModel.create_all emits
    them in the order the model declares them, so two identical schemas
    have different SQL. Column order is left out for the same reason.
    """
    insp = inspect(create_engine(f"sqlite:///{path}"))
    shape = {}
    for table in sorted(insp.get_table_names()):
        if table == "alembic_version":
            continue
        shape[table] = {
            "columns": {c["name"]: (str(c["type"]), c["nullable"]) for c in insp.get_columns(table)},
            "pk": tuple(sorted(insp.get_pk_constraint(table)["constrained_columns"])),
            "fks": sorted((tuple(f["constrained_columns"]), f["referred_table"],
                           tuple(f["referred_columns"])) for f in insp.get_foreign_keys(table)),
            "unique": sorted((u["name"], tuple(u["column_names"]))
                             for u in insp.get_unique_constraints(table)),
            "indexes": sorted((i["name"], tuple(i["column_names"]), bool(i["unique"]))
                              for i in insp.get_indexes(table)),
        }
    return shape


def _drift(target_engine):
    with target_engine.connect() as conn:
        return compare_metadata(MigrationContext.configure(conn), SQLModel.metadata)


def test_migrations_build_what_the_models_describe():
    """A new install gets its schema from the migrations, not from
    create_all - so the migrations have to produce exactly what create_all
    would have. If they ever stop agreeing, a fresh farm quietly gets a
    different database from an upgraded one, and the difference shows up as
    a "no such column" months later on whichever of the two nobody tested."""
    tmp = tempfile.mkdtemp()
    try:
        by_migration = os.path.join(tmp, "migrated.db")
        by_models = os.path.join(tmp, "created.db")
        run_migrations(create_engine(f"sqlite:///{by_migration}"), snapshot=False)
        SQLModel.metadata.create_all(create_engine(f"sqlite:///{by_models}"))

        migrated, created = _table_shape(by_migration), _table_shape(by_models)
        assert set(migrated) == set(created), (
            f"tables only one way builds: {set(migrated) ^ set(created)}")
        for table in migrated:
            assert migrated[table] == created[table], (
                f"{table} differs\n  migrations: {migrated[table]}\n"
                f"  models:     {created[table]}")
        assert not _drift(create_engine(f"sqlite:///{by_migration}")), (
            "a database built from the migrations does not match models.py - "
            "a migration is missing")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_a_database_from_before_migrations_is_caught_up_and_stamped():
    """The upgrade path every existing farm takes exactly once.

    Built here the way those databases were built - at the BASELINE shape,
    which is what a database that predates Alembic actually has, not
    today's - then damaged the way a farm several releases behind is
    damaged, a table and a column it never received. It has to come out at
    the newest revision, with the post-baseline migrations actually applied
    and its rows untouched. A farm that skipped four releases is the case
    this has to survive, not a farm that updates weekly.
    """
    tmp = tempfile.mkdtemp()
    try:
        path = os.path.join(tmp, "old.db")
        old_engine = create_engine(f"sqlite:///{path}")
        legacy_schema_catch_up(old_engine, _baseline_database())
        with Session(old_engine) as s:
            s.add(Block(id="15", name="Blok 15", variety="Mauritius", trees=100, hectares=1.5))
            s.add(Worker(id="001", first_name="Thandi", last_name="N", name="Thandi N"))
            s.commit()

        con = sqlite3.connect(path)
        con.execute("DROP TABLE setupstate")
        con.execute("ALTER TABLE adminuser DROP COLUMN must_change_password")
        con.commit()
        con.close()

        assert current_revision(old_engine) is None, "a pre-Alembic database claimed a revision"
        run_migrations(old_engine, snapshot=False)

        assert current_revision(old_engine) == head_revision(), (
            "an existing farm database did not end up at the newest revision")
        insp = inspect(old_engine)
        assert insp.has_table("setupstate"), "a table the farm never received was not restored"
        assert "must_change_password" in {c["name"] for c in insp.get_columns("adminuser")}, \
            "a column the farm never received was not restored"
        assert not _drift(old_engine), (
            "an adopted database does not match models.py once it is up to date")
        with Session(old_engine) as s:
            assert s.exec(select(func.count()).select_from(Block)).one() == 1
            assert s.exec(select(func.count()).select_from(Worker)).one() == 1
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_the_catch_up_lands_exactly_on_the_baseline():
    """The catch-up is what a pre-Alembic farm gets instead of the baseline
    migration, and the database is stamped at the baseline immediately
    afterwards - so it has to arrive at the same schema the baseline
    migration would have built. Not a subset, not a superset.

    A superset is the failure that actually happened: while the catch-up
    worked off models.py it built TODAY's schema, and the first migration
    after the baseline then tried to add a column that was already there.
    That kills the update on exactly the farms furthest behind, and it
    could not happen until a migration finally added a column - which is to
    say, it would have been found on a farm rather than here.
    """
    tmp = tempfile.mkdtemp()
    try:
        baseline_path = os.path.join(tmp, "baseline.db")
        command.upgrade(_config(create_engine(f"sqlite:///{baseline_path}")), BASELINE_REVISION)

        caught_up_path = os.path.join(tmp, "caught_up.db")
        legacy_schema_catch_up(create_engine(f"sqlite:///{caught_up_path}"), _baseline_database())

        caught_up, baseline = _table_shape(caught_up_path), _table_shape(baseline_path)
        assert set(caught_up) == set(baseline), (
            f"tables only one of the two builds: {set(caught_up) ^ set(baseline)}")
        for table in baseline:
            assert caught_up[table] == baseline[table], (
                f"{table} differs\n  catch-up: {caught_up[table]}\n"
                f"  baseline: {baseline[table]}")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_migrating_an_up_to_date_database_changes_nothing():
    """Startup runs migrations every time. All but the first must be a no-op."""
    tmp = tempfile.mkdtemp()
    try:
        path = os.path.join(tmp, "current.db")
        run_migrations(create_engine(f"sqlite:///{path}"), snapshot=False)
        before = _table_shape(path)
        run_migrations(create_engine(f"sqlite:///{path}"), snapshot=False)
        run_migrations(create_engine(f"sqlite:///{path}"), snapshot=False)
        assert _table_shape(path) == before, "re-running migrations altered the schema"
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def test_baseline_is_still_the_root_revision():
    """BASELINE_REVISION is what pre-Alembic farms get stamped at. Pointed at
    anything other than the root, it would stamp an old farm as though
    migrations it never received had already run - and those migrations
    would then never run, silently."""
    script = ScriptDirectory.from_config(_config(engine))
    baseline = script.get_revision(BASELINE_REVISION)
    assert baseline is not None, f"{BASELINE_REVISION} is not in migrations/versions"
    assert baseline.down_revision is None, (
        f"the baseline {BASELINE_REVISION} is no longer the root revision "
        f"(it now follows {baseline.down_revision})")
    assert len(script.get_heads()) == 1, (
        f"the migration history has branched: {script.get_heads()} - farms would apply "
        f"whichever head Alembic happened to pick")


def test_pre_migration_snapshot_is_a_faithful_full_copy():
    """The copy taken before a migration is a rollback point, so it keeps the
    weather history the nightly archive deliberately throws away - a restore
    can re-download 42 MB of weather from Open-Meteo, a rollback at the wrong
    moment on an update night cannot."""
    tmp = tempfile.mkdtemp()
    real_db, real_backups = backup.DB_PATH, backup.BACKUPS_DIR
    try:
        path = os.path.join(tmp, "farm.db")
        backup.DB_PATH = path
        backup.BACKUPS_DIR = os.path.join(tmp, "backups")
        os.makedirs(backup.BACKUPS_DIR)

        temp_engine = create_engine(f"sqlite:///{path}")
        run_migrations(temp_engine, snapshot=False)  # any populated database will do here
        with Session(temp_engine) as s:
            s.add(Block(id="15", name="Blok 15", variety="Mauritius", trees=100, hectares=1.5))
            for hour in range(48):
                s.add(WeatherHistory(timestamp=datetime(2025, 1, 1 + hour // 24, hour % 24),
                                      temp_c=20.0 + hour))
            s.commit()

        copy = backup.snapshot_before_migration("selftest")
        con = sqlite3.connect(copy)
        try:
            assert con.execute("SELECT count(*) FROM weatherhistory").fetchone()[0] == 48, \
                "the pre-migration copy dropped the weather history"
            assert con.execute("SELECT count(*) FROM block").fetchone()[0] == 1
        finally:
            con.close()

        for i in range(5):
            backup.snapshot_before_migration(f"selftest{i}")
        kept = backup._pre_migration_filenames()
        assert len(kept) == backup.PRE_MIGRATION_KEEP, f"retention kept {len(kept)}: {kept}"
        assert backup._backup_filenames() == [], (
            "the nightly pruner can see the pre-migration copies - it would delete "
            "rollback points to make room for backups")
    finally:
        backup.DB_PATH, backup.BACKUPS_DIR = real_db, real_backups
        shutil.rmtree(tmp, ignore_errors=True)


def test_this_server_is_at_the_newest_revision():
    if not os.path.exists(DB_PATH):
        raise Skip("no database yet - start the server once")
    if not inspect(engine).has_table("alembic_version"):
        raise Skip("this database has not been migrated yet - start the server against "
                    "it once, which stamps it at the baseline, then re-run")
    current, head = current_revision(), head_revision()
    assert current == head, (
        f"this database is at {current} but the code here is at {head} - the server "
        f"has not been restarted since the update, so it is running new code against "
        f"an old schema")


def test_this_server_matches_the_models():
    """Read-only drift check on the live database. The same comparison
    migrate.py prints at startup, asserted here so it cannot scroll past
    unnoticed in a Scheduled Task's console."""
    if not inspect(engine).has_table("alembic_version"):
        raise Skip("this database has not been migrated yet")
    diff = _drift(engine)
    assert not diff, "the live schema does not match models.py: " + "; ".join(
        str(entry) for entry in diff)


# ---------------------------------------------------------------------------
# Master data imports
# ---------------------------------------------------------------------------
def test_replacing_all_blocks_with_an_empty_file_is_refused():
    """"Replace all" reads the uploaded file as the farm's new complete block
    list, so an empty one used to mean "the new list is empty" and retired
    every block the farm had - reporting {"imported": 0, "deactivated": 21},
    which reads as success. The template the wizard hands out is itself a
    headings-only file, so the mistake is one download and one upload away."""
    tmp = tempfile.mkdtemp()
    try:
        path = os.path.join(tmp, "blocks.db")
        temp_engine = create_engine(f"sqlite:///{path}")
        run_migrations(temp_engine, snapshot=False)
        with Session(temp_engine) as s:
            s.add(Block(id="15", name="Blok 15", variety="Mauritius", trees=100, hectares=1.5))
            s.add(Block(id="16", name="Blok 16", variety="Mauritius", trees=80, hectares=1.2))
            s.commit()

        headings_only = UploadFile(filename="blocks.csv",
                                    file=io.BytesIO(b"id,name,variety,trees,hectares,active\n"))
        with Session(temp_engine) as s:
            try:
                asyncio.run(import_blocks(file=headings_only, replace=True, session=s, admin=None))
            except HTTPException as e:
                assert e.status_code == 400, e.status_code
                assert "no data rows" in e.detail
            else:
                raise AssertionError("an empty file was accepted as the farm's new block list")

        with Session(temp_engine) as s:
            active = s.exec(select(func.count()).select_from(Block).where(Block.active == True)).one()  # noqa: E712
        assert active == 2, f"{2 - active} block(s) were retired by a file with nothing in it"
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


# ---------------------------------------------------------------------------
# First-run setup wizard state
#
# The one thing worth asserting against a live database: that this farm is
# NOT offered the wizard. Getting `required` wrong in that direction sits an
# established farm in front of an empty setup form instead of its own
# dashboard, and the check that prevents it deliberately does not trust a
# single flag - so it is worth testing on a real database rather than only
# on a contrived one.
# ---------------------------------------------------------------------------
def _require_setup_table():
    """SetupState is created at server startup (main.on_startup ->
    run_migrations). A database no server has opened since this
    release simply does not have it yet - that is not a failure, and on a
    live farm server, which by definition has started, it never happens."""
    if not inspect(engine).has_table("setupstate"):
        raise Skip("no setupstate table yet - start the server against this "
                    "database once, then re-run")


def test_setup_not_required_on_a_configured_farm():
    _require_setup_table()
    with Session(engine) as s:
        state = build_setup_state(s)
        settings = s.exec(select(SystemSetting)).first()

    if state["started_at"] and not state["completed_at"]:
        # The wizard is open and unfinished - on this farm, right now,
        # somebody is part way through it. Still being offered is correct
        # here, and is the whole reason started_at exists: step 1 writes the
        # farm name, so without it the remaining steps become unreachable on
        # the next page load. Assert that rather than reading a half-done
        # setup as a failure.
        assert state["required"] is True, "an unfinished setup wizard stopped being offered"
        return

    named = bool(settings and (settings.farm_name or "").strip())
    picked = state["harvest_records"] > 0
    if not (named or picked or state["completed_at"]):
        return  # a genuinely blank database - the wizard SHOULD be offered
    assert state["required"] is False, (
        f"this farm would be sent through the setup wizard "
        f"(named={named}, crates={state['harvest_records']}, "
        f"started={state['started_at']}, completed={state['completed_at']})")


def test_setup_state_reports_every_step():
    """Every step the wizard knows about has to come back with a verdict -
    a missing key reads as "not done" in the browser and silently reopens
    a step the farm has already finished."""
    _require_setup_table()
    with Session(engine) as s:
        state = build_setup_state(s)
    expected = {"identity", "location", "rate", "thresholds", "blocks",
                "workers", "devices", "history"}
    assert set(state["steps"]) == expected, f"steps drifted: {set(state['steps'])}"
    for key, step in state["steps"].items():
        assert isinstance(step.get("done"), bool), f"{key} has no boolean 'done'"


def test_setup_state_writes_nothing():
    """build_setup_state only reads. It runs on every admin page load, and
    it must not be what creates the SetupState row - the absence of that row
    is exactly how a database that predates the wizard is recognised."""
    _require_setup_table()
    with Session(engine) as s:
        before = s.exec(select(func.count()).select_from(SetupState)).one()
    with Session(engine) as s:
        build_setup_state(s)
    with Session(engine) as s:
        after = s.exec(select(func.count()).select_from(SetupState)).one()
    assert before == after, f"SetupState rows changed: {before} -> {after}"


# ---------------------------------------------------------------------------
# Report workbook
# ---------------------------------------------------------------------------
def _report_workbook():
    with Session(engine) as s:
        resp = historical_harvest_data_report(session=s, admin=None)
    return openpyxl.load_workbook(io.BytesIO(resp.body)), resp


def test_report_has_expected_sheets():
    wb, _ = _report_workbook()
    for name in ("Blocks", "Notes", "Season Summary", "Block by Year"):
        assert name in wb.sheetnames, f"missing sheet {name}"


def test_report_block_by_year_reconciles_with_season_summary():
    """The two cross-era sheets are built from the same figures by different
    routes; if they disagree, one of them is wrong."""
    wb, _ = _report_workbook()
    ss = {r[0]: r[1] for r in wb["Season Summary"].iter_rows(min_row=2, values_only=True)}
    rows = list(wb["Block by Year"].iter_rows(values_only=True))
    header, total_row = rows[0], rows[-1]
    assert total_row[0] == "TOTAL"
    for i, h in enumerate(header):
        if isinstance(h, str) and h.isdigit():
            year = int(h)
            assert abs((total_row[i] or 0) - (ss.get(year) or 0)) < 0.5, \
                f"{year}: grid {total_row[i]} vs summary {ss.get(year)}"


def test_report_block_by_year_row_totals():
    wb, _ = _report_workbook()
    rows = list(wb["Block by Year"].iter_rows(values_only=True))
    header = rows[0]
    year_cols = [i for i, h in enumerate(header) if isinstance(h, str) and h.isdigit()]
    for row in rows[1:]:
        got = sum(row[i] or 0 for i in year_cols)
        assert abs(got - (row[-1] or 0)) < 0.5, f"{row[0]} row total {row[-1]} != {got}"


def test_report_season_summary_matches_database():
    wb, _ = _report_workbook()
    ss = {r[0]: (r[1], r[4]) for r in wb["Season Summary"].iter_rows(min_row=2, values_only=True)}
    with Session(engine) as s:
        daily = defaultdict(float)
        for h in s.exec(select(HistoricalHarvest)).all():
            daily[h.season_year] += h.kg
        annual = defaultdict(float)
        for a in s.exec(select(HistoricalAnnualYield)).all():
            annual[a.season_year] += a.kg
    for year, kg in daily.items():
        assert abs(ss[year][0] - kg) < 0.5, f"{year}: report {ss[year][0]} vs table {kg}"
        assert ss[year][1].startswith("Daily"), ss[year]
    for year, kg in annual.items():
        assert abs(ss[year][0] - kg) < 0.5, f"{year}: report {ss[year][0]} vs table {kg}"
        assert ss[year][1].startswith("Season total"), ss[year]


def test_report_year_on_year_only_for_consecutive_seasons():
    """The record has no 2010-2011; a jump across that gap must not be
    presented as a year-on-year change."""
    wb, _ = _report_workbook()
    rows = list(wb["Season Summary"].iter_rows(min_row=2, values_only=True))
    totals = {r[0]: r[1] for r in rows}
    for year, total, change, _blocks, _gran in rows:
        if year - 1 not in totals:
            assert change is None, f"{year} shows a change with no {year - 1} to compare against"
        elif totals[year - 1]:
            expected = (total - totals[year - 1]) / totals[year - 1]
            assert abs(change - expected) < 1e-9, year


def test_report_filename_spans_whole_workbook():
    wb, resp = _report_workbook()
    disposition = resp.headers.get("content-disposition", "")
    years = [int(n) for n in wb["Season Summary"].iter_rows(min_row=2, max_col=1, values_only=True)
             for n in n if n is not None]
    assert f"{min(years)}_{max(years)}" in disposition, disposition


def test_report_per_season_sheets_match_daily_table():
    wb, _ = _report_workbook()
    with Session(engine) as s:
        daily = defaultdict(float)
        for h in s.exec(select(HistoricalHarvest)).all():
            daily[h.season_year] += h.kg
    for year, kg in daily.items():
        ws = wb[str(year)]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        # last row is a formula footer; sum the data rows' own Total column
        total = sum(r[-1] or 0 for r in rows if isinstance(r[0], (date, datetime)))
        assert abs(total - kg) < 1.0, f"{year} sheet totals {total} vs table {kg}"


# ---------------------------------------------------------------------------
# Other endpoints still build (regression guard for the shared weather load)
# ---------------------------------------------------------------------------
def test_analysis_and_weather_endpoints_build():
    with Session(engine) as s:
        a = build_analysis_summary(s)
        w = build_weather_history(s)
    assert a["historical_years"], "analysis lost its historical seasons"
    assert w["points"], "weather history built no points"
    assert w["years"] == sorted(set(w["years"]))


def test_analysis_unaffected_by_annual_import():
    """The Analysis tab is deliberately daily-only; the annual-totals import
    must not have leaked its season-only years into it."""
    with Session(engine) as s:
        a = build_analysis_summary(s)
        annual = set(s.exec(select(HistoricalAnnualYield.season_year).distinct()).all())
    assert not (set(a["historical_years"]) & annual), \
        "season-total years leaked into the daily Analysis tab"


def test_current_season_never_scored_as_final():
    with Session(engine) as s:
        settings = s.exec(select(SystemSetting)).first()
        r = build_risk_summary(s)
    current = settings.current_harvest_year if settings else date.today().year
    season = next(x for x in r["seasons"] if x["year"] == current)
    assert season["is_current"] is True
    if season["known_count"] < len(DRIVERS):
        assert season["risk_score"] is None, "partial season reported a final score"


def main():
    print("Boord self-test")
    print("=" * 60)

    section("Pure functions: scoring")
    for fn in (test_risk_points_scaling, test_risk_points_clamps_beyond_history,
               test_risk_points_degenerate, test_band_boundaries, test_ols_fit,
               test_reference_label):
        check(fn.__name__, fn)

    section("Pure functions: driver aggregation and windowing")
    for fn in (test_driver_value_aggregations, test_driver_value_missing_data,
               test_every_driver_agg_is_implemented, test_window_status_transitions,
               test_segment_days_partition, test_segment_days_zero_horizon_has_no_forecast,
               test_project_driver_intensive_vs_extensive,
               test_project_driver_data_gap_falls_back):
        check(fn.__name__, fn)

    section("Driver configuration")
    for fn in (test_driver_definitions_wellformed, test_driver_fields_are_fetchable_for_all_eras,
               test_regression_start_not_before_reference_start,
               test_driver_values_are_physically_plausible,
               test_averaging_drivers_blend_by_day_weight,
               test_regression_seasons_all_had_a_bearing_orchard):
        check(fn.__name__, fn)

    section("Stored data integrity")
    for fn in (test_weather_history_continuous, test_no_duplicate_weather_hours,
               test_weather_covers_every_reference_window,
               test_historical_tables_dont_overlap, test_annual_yield_block_ids_resolve):
        check(fn.__name__, fn)

    section("Risk indicator")
    for fn in (test_risk_summary_every_reference_season_scored,
               test_risk_component_points_sum_to_score,
               test_risk_reference_extremes_hit_the_rails,
               test_reference_years_are_modern_orchard_only,
               test_risk_and_forecast_agree_on_shared_state,
               test_current_season_never_scored_as_final):
        check(fn.__name__, fn)

    section("Harvest forecast")
    for fn in (test_forecast_scenarios_ordered_and_bounded,
               test_forecast_predictions_within_recorded_range,
               test_forecast_labels_match_their_own_ranges,
               test_forecast_degrades_without_network,
               test_selftest_writes_nothing):
        check(fn.__name__, fn)

    section("Historical Harvest Data report")
    for fn in (test_report_has_expected_sheets,
               test_report_block_by_year_reconciles_with_season_summary,
               test_report_block_by_year_row_totals,
               test_report_season_summary_matches_database,
               test_report_year_on_year_only_for_consecutive_seasons,
               test_report_filename_spans_whole_workbook,
               test_report_per_season_sheets_match_daily_table):
        check(fn.__name__, fn)

    section("Schema migrations")
    for fn in (test_migrations_build_what_the_models_describe,
               test_a_database_from_before_migrations_is_caught_up_and_stamped,
               test_the_catch_up_lands_exactly_on_the_baseline,
               test_migrating_an_up_to_date_database_changes_nothing,
               test_baseline_is_still_the_root_revision,
               test_pre_migration_snapshot_is_a_faithful_full_copy,
               test_this_server_is_at_the_newest_revision,
               test_this_server_matches_the_models):
        check(fn.__name__, fn)

    section("Master data imports")
    check(test_replacing_all_blocks_with_an_empty_file_is_refused.__name__,
          test_replacing_all_blocks_with_an_empty_file_is_refused)

    section("Setup state")
    for fn in (test_setup_not_required_on_a_configured_farm,
               test_setup_state_reports_every_step,
               test_setup_state_writes_nothing):
        check(fn.__name__, fn)

    section("Other endpoints")
    for fn in (test_analysis_and_weather_endpoints_build,
               test_analysis_unaffected_by_annual_import):
        check(fn.__name__, fn)

    print("\n" + "=" * 60)
    if _skipped:
        print(f"SKIPPED: {len(_skipped)}")
        for name, why in _skipped:
            print(f"  - {name}: {why}")
    if _failed:
        print(f"FAILED: {len(_failed)} of {_passed + len(_failed)} checks")
        for name, exc in _failed:
            print(f"  - {name}: {type(exc).__name__}: {exc}")
        return 1
    print(f"PASSED: all {_passed} checks")
    return 0


if __name__ == "__main__":
    sys.exit(main())
